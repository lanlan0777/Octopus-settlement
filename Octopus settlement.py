import os
import shutil
import time
import datetime
#from datetime import datetime
from dateutil.parser import parse
import zipfile
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment

last_process_time = None

onedrive_folder_path = "C:\\Users\\lcyer\\OneDrive - Enjoyable Way Limited\\New folder\\"
dest_zip_folder = "C:\\Users\\lcyer\\EW-Group Dropbox\\Company-Other\\Property\Ricky Centre CP(RCCP) - EW\\POS settlement\\Octopus archive\\zip file\\"
dest_unzip_folder = "C:\\Users\\lcyer\\EW-Group Dropbox\\Company-Other\\Property\Ricky Centre CP(RCCP) - EW\\POS settlement\\Octopus archive\\"
time_log_file_name = "C:\\Users\\lcyer\\OneDrive\\Documents\\octopus settlement process time.txt"

SETTLEMENT_DATE = 0
NET_VALUE = 16
NET_CHARGES = 18
NET_ENTITLEMENT = 19
Octopus_Record = []
password = 'EFT5210'
'''---------------------------------------------------------------------
Function: Get the last runtime of the program from txt file
Params:
    None
Return:
    None
---------------------------------------------------------------------'''   
def get_last_process_time():
    global last_process_time
    with open(time_log_file_name, "r") as f:  # 打开文件
        last_process_time = parse(f.read())  # 读取文件
        print("last_process_time is %s:" %last_process_time)
        
'''---------------------------------------------------------------------
Function: Check if the file is changed sicne the last time the program was run
Params:
    client_modified: For files, this is the modification time set by the desktop client when the file was added to Dropbox
    server_modified: The last time the file was modified on Dropbox
Return:
    True: The file is new added or updated since the last time the program was run
    False: The file is not changed sicne the last time the program was run
---------------------------------------------------------------------'''
def is_file_updated(file_name):
    ctime = time.gmtime(os.path.getmtime(onedrive_folder_path +file_name))  
    str_ctime = parse(time.strftime("%Y-%m-%d %H:%M:%S", ctime))

    if str_ctime > last_process_time:  
        print("new settlement:",file_name)
        return True
    else:
        #print("old receipt")
        return False
    
'''---------------------------------------------------------------------
Function: record this runtime of the program to txt file
Params:
    None
Return:
    None
---------------------------------------------------------------------''' 
def write_process_time():
    with open(time_log_file_name,"w") as f:
        f.write(str(datetime.datetime.now()+datetime.timedelta(hours=-8)))  
        
get_last_process_time()

onedrive_zip_list = os.listdir(onedrive_folder_path)

for each in onedrive_zip_list:
    if is_file_updated(each):
        if ".zip" in each:
            shutil.copy(onedrive_folder_path+each,dest_zip_folder+each)  #copy zip file to dropbox
            zf = zipfile.ZipFile(onedrive_folder_path+each,"r")
            zf.setpassword(pwd=password.encode())
            zf.extractall(dest_unzip_folder+each[0:-4])                  #unzip folder to dropbox
           
            #extract octopus data from SET051ShiftOOS.CSV
            f = pd.read_csv(dest_unzip_folder+each[0:-4]+"\\SET051ShiftOOS.CSV")
            row_list = f.values.tolist()

            Octopus_Record.append(row_list[0][SETTLEMENT_DATE].replace("-","/"))
            Octopus_Record.append(row_list[0][NET_VALUE])
            if isinstance(row_list[0][NET_CHARGES], str):
                Octopus_Record.append(float("-" + row_list[0][NET_CHARGES][1:-1]))
            else:
                Octopus_Record.append(row_list[0][NET_CHARGES])
            Octopus_Record.append(row_list[0][NET_ENTITLEMENT])
            print(Octopus_Record)
            
            #write octopus data to O_summary.xlsx               
            wb = openpyxl.load_workbook(dest_unzip_folder+"O_summary.xlsx")
            ws = wb.active
            ws.append(Octopus_Record)
            last_row = ws.max_row
            print(last_row)
            ws.cell(last_row,1).number_format ="dd/mm/yyyy" 
            ws.cell(last_row,2).number_format ='#,##0.00'
            ws.cell(last_row,3).number_format ='#,##0.00' 
            ws.cell(last_row,4).number_format ='#,##0.00' 
            alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(last_row,1).alignment = alignment
            ws.cell(last_row,2).alignment = alignment
            ws.cell(last_row,3).alignment = alignment
            ws.cell(last_row,4).alignment = alignment
            wb.save(filename=dest_unzip_folder+"O_summary.xlsx")
            
            Octopus_Record.clear()

write_process_time()


